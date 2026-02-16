"""
Unit-тесты для финансовой точности конвертера.
Запуск: pytest test_app.py -v
"""
import pytest
import pandas as pd
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from openpyxl import load_workbook

# Импортируем тестируемые функции напрямую
from app import parse_amount, norm_text


# ============================================================
# 1. Парсинг форматов
# ============================================================

class TestParseAmountFormats:
    """Тесты для различных числовых форматов."""

    def test_swiss_apostrophe(self):
        """Швейцарский формат: 1'000.00"""
        val, reason, _ = parse_amount("1'000.00")
        assert reason is None
        assert val == Decimal("1000.00")

    def test_swiss_apostrophe_typographic(self):
        """Типографский апостроф: 1\u2019000.00"""
        val, reason, _ = parse_amount("1\u2019000.00")
        assert reason is None
        assert val == Decimal("1000.00")

    def test_accounting_parentheses(self):
        """Бухгалтерские скобки: (500) = -500"""
        val, reason, _ = parse_amount("(500)")
        assert reason is None
        assert val == Decimal("-500")

    def test_accounting_parentheses_with_decimals(self):
        """Бухгалтерские скобки с дробной частью: (1,234.56) = -1234.56"""
        val, reason, _ = parse_amount("(1,234.56)")
        assert reason is None
        assert val == Decimal("-1234.56")

    def test_eu_format(self):
        """Европейский формат: 1.234,56"""
        val, reason, _ = parse_amount("1.234,56")
        assert reason is None
        assert val == Decimal("1234.56")

    def test_us_format(self):
        """Американский формат: 1,234.56"""
        val, reason, _ = parse_amount("1,234.56")
        assert reason is None
        assert val == Decimal("1234.56")

    def test_space_as_thousands_separator(self):
        """Пробел как разделитель тысяч: 1 000"""
        val, reason, _ = parse_amount("1 000")
        assert reason is None
        assert val == Decimal("1000")

    def test_nbsp_as_thousands_separator(self):
        """NBSP как разделитель тысяч: 1\u00A0000,50"""
        val, reason, _ = parse_amount("1\u00A0000,50")
        assert reason is None
        assert val == Decimal("1000050") or val == Decimal("1000.50")
        # после удаления NBSP → "1000,50" → Decimal("1000.50")
        assert val == Decimal("1000.50")

    def test_trailing_minus(self):
        """Постфиксный минус: 500-"""
        val, reason, _ = parse_amount("500-")
        assert reason is None
        assert val == Decimal("-500")

    def test_currency_symbol_rub(self):
        """Символ валюты: 1000 ₽"""
        val, reason, _ = parse_amount("1000 ₽")
        assert reason is None
        assert val == Decimal("1000")

    def test_currency_symbol_usd(self):
        """Символ валюты: $1,000.00"""
        val, reason, _ = parse_amount("$1,000.00")
        assert reason is None
        assert val == Decimal("1000.00")

    def test_already_numeric_int(self):
        """Числовой int."""
        val, reason, _ = parse_amount(42)
        assert reason is None
        assert val == Decimal("42")

    def test_already_numeric_float(self):
        """Числовой float → Decimal через str."""
        val, reason, _ = parse_amount(0.1)
        assert reason is None
        assert isinstance(val, Decimal)
        # Decimal(str(0.1)) = Decimal('0.1'), а не Decimal(0.1) = Decimal('0.1000...00005...')
        assert val == Decimal("0.1")

    def test_em_dash_minus(self):
        """Длинное тире как минус."""
        val, reason, _ = parse_amount("—500")
        assert reason is None
        assert val == Decimal("-500")


# ============================================================
# 2. Пустые / невалидные значения
# ============================================================

class TestParseAmountRejects:
    """Тесты для значений, которые НЕ должны стать суммами."""

    def test_none(self):
        val, reason, _ = parse_amount(None)
        assert val is None
        assert reason == "EMPTY"

    def test_nan(self):
        val, reason, _ = parse_amount(float("nan"))
        assert val is None
        assert reason == "EMPTY"

    def test_empty_string(self):
        val, reason, _ = parse_amount("")
        assert val is None
        assert reason == "EMPTY"

    def test_dash(self):
        val, reason, _ = parse_amount("-")
        assert val is None
        assert reason == "EMPTY"

    def test_bool_marker_da(self):
        val, reason, _ = parse_amount("Да")
        assert val is None
        assert reason == "NON_AMOUNT"

    def test_bool_marker_net(self):
        val, reason, _ = parse_amount("Нет")
        assert val is None
        assert reason == "NON_AMOUNT"

    def test_pure_text(self):
        """Чистый текст без цифр."""
        val, reason, _ = parse_amount("текст")
        assert val is None
        assert reason == "NON_AMOUNT"

    def test_text_with_units(self):
        """Текст без цифр: 'шт'."""
        val, reason, _ = parse_amount("шт")
        assert val is None
        assert reason == "NON_AMOUNT"


# ============================================================
# 3. Decimal-точность при массовом суммировании
# ============================================================

class TestDecimalPrecision:
    """Тесты, доказывающие что Decimal не теряет копейки."""

    def test_classic_float_trap(self):
        """0.1 + 0.2 == 0.3 — работает в Decimal, ломается в float."""
        a = Decimal("0.1")
        b = Decimal("0.2")
        assert a + b == Decimal("0.3")
        # для контраста — float
        assert 0.1 + 0.2 != 0.3  # float-артефакт

    def test_10k_small_transactions(self):
        """
        10 000 транзакций по 0.0001.
        Ожидание: 10000 * 0.0001 = 1.0000 ровно.
        Float даст ~0.9999999999999062.
        """
        tx_value = Decimal("0.0001")
        total = sum(tx_value for _ in range(10_000))
        assert total == Decimal("1.0000")

    def test_10k_penny_transactions(self):
        """
        10 000 транзакций по 0.01.
        Ожидание: 100.00 ровно.
        """
        tx_value = Decimal("0.01")
        total = sum(tx_value for _ in range(10_000))
        assert total == Decimal("100.00")

    def test_parsed_amounts_sum(self):
        """
        Парсим 10 000 значений "0.01" через parse_amount и суммируем.
        Результат должен быть Decimal("100.00").
        """
        values = []
        for _ in range(10_000):
            val, reason, _ = parse_amount("0.01")
            assert reason is None
            values.append(val)

        total = sum(values, Decimal(0))
        assert total == Decimal("100.00"), f"Expected 100.00, got {total}"

    def test_parsed_amounts_sum_mixed_formats(self):
        """
        Суммирование разных форматов: результат точный.
        """
        inputs = [
            "1'000.50",   # 1000.50
            "(200,25)",   # -200.25  (скобки + EU-запятая)
            "300.00",     # 300.00
            "50-",        # -50.00
        ]
        expected = Decimal("1000.50") - Decimal("200.25") + Decimal("300.00") - Decimal("50")
        # = 1050.25

        total = Decimal(0)
        for inp in inputs:
            val, reason, _ = parse_amount(inp)
            assert reason is None, f"Failed to parse: {inp}"
            total += val

        assert total == expected, f"Expected {expected}, got {total}"


# ============================================================
# 4. Невалидные данные не попадают в итог
# ============================================================

class TestInvalidDataIsolation:
    """Проверяем, что невалидные строки НИКОГДА не попадают в result_rows."""

    def test_text_in_amount_column_goes_to_error(self):
        """
        Создаём DataFrame как если бы 'текст' попал в колонку с суммами.
        parse_amount должен вернуть NON_AMOUNT, и строка не должна стать операцией.
        """
        test_values = ["текст", "abc", "N/A", "---", "!!!", "Оплата чего-то"]
        for tv in test_values:
            val, reason, dbg = parse_amount(tv)
            assert val is None, f"'{tv}' was parsed as {val}, expected None"
            assert reason in {"NON_AMOUNT", "EMPTY", "PARSE_FAIL"}, \
                f"'{tv}' reason={reason}, expected rejection"

    def test_mixed_text_and_number_returns_number(self):
        """'1000 руб.' → должен распарсить как 1000."""
        val, reason, _ = parse_amount("1000 руб.")
        assert reason is None
        assert val == Decimal("1000")

    def test_only_currency_symbol(self):
        """Только '₽' без цифр → NON_AMOUNT."""
        val, reason, _ = parse_amount("₽")
        assert val is None
        assert reason == "NON_AMOUNT"


# ============================================================
# 5. Roundtrip: parse → quantize → sum → Excel
# ============================================================

class TestExcelRoundtrip:
    """Проверяем, что данные корректно добираются до Excel."""

    def test_decimal_to_excel_and_back(self):
        """
        Записываем Decimal-суммы через openpyxl и читаем обратно.
        Проверяем, что значения сохранены точно.
        """
        _two_places = Decimal("0.01")

        # Генерируем 100 сумм по 0.01 и ожидаем total = 1.00
        values = [Decimal("0.01").quantize(_two_places, rounding=ROUND_HALF_UP) for _ in range(100)]
        total_expected = sum(values, Decimal(0))

        # Записываем в Excel
        df = pd.DataFrame({
            "Приход": [float(v) for v in values],  # openpyxl не пишет Decimal, конвертим в float
        })

        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Test")
        buf.seek(0)

        # Читаем обратно
        df_read = pd.read_excel(buf, sheet_name="Test")
        total_read = Decimal(str(df_read["Приход"].sum()))

        # Допуск: 0.01 (1 копейка) на 100 строк
        assert abs(total_read - total_expected) < Decimal("0.01"), \
            f"Excel roundtrip error: expected {total_expected}, got {total_read}"

    def test_zero_hidden_format_preserves_numeric_type(self):
        """
        Записываем 0 с кастомным форматом — ячейка остаётся числовой.
        """
        buf = BytesIO()
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Приход"
        ws["A2"] = 0
        ws["A2"].number_format = '#,##0.00;-#,##0.00;""'
        ws["A3"] = 100.50
        ws["A3"].number_format = '#,##0.00;-#,##0.00;""'
        wb.save(buf)
        buf.seek(0)

        # Читаем обратно
        wb2 = load_workbook(buf)
        ws2 = wb2.active
        assert ws2["A2"].value == 0  # Числовой 0, не строка
        assert isinstance(ws2["A2"].value, (int, float))
        assert ws2["A3"].value == 100.50
        assert isinstance(ws2["A3"].value, float)


# ============================================================
# 6. Norm_text — обработка экзотических пробелов
# ============================================================

class TestNormText:
    """Проверяем нормализацию текста."""

    def test_nbsp(self):
        assert norm_text("hello\u00A0world") == "hello world"

    def test_narrow_nbsp(self):
        assert norm_text("1\u202F000") == "1 000"

    def test_thin_space(self):
        assert norm_text("1\u2009000") == "1 000"

    def test_multiple_spaces(self):
        assert norm_text("  a   b  ") == "a b"

    def test_none(self):
        assert norm_text(None) == ""
